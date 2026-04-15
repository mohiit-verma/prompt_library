"""
Feedback persistence layer.

Ratings, comments, and filled-prompt history are written to a dedicated
feedback.xlsx file, intentionally decoupled from the prompt source file.
This means uploading a new prompt library never wipes existing user feedback.
"""
from __future__ import annotations

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import pandas as pd

from config import APP_CONFIG, AppConfig

logger = logging.getLogger("prompt_library.feedback")


# ── Sheet bootstrap ───────────────────────────────────────────────────────────

def ensure_feedback_file(
    path: str, config: AppConfig = APP_CONFIG
) -> None:
    """Creates the feedback workbook and required sheets if they do not exist."""
    wb_path = Path(path)

    if wb_path.exists():
        wb = openpyxl.load_workbook(wb_path)
    else:
        logger.info("Creating new feedback file: %s", wb_path.resolve())
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    _ensure_sheet(
        wb, config.ratings_sheet,
        ["Prompt ID", "Rating", "Session ID", "Created At"],
    )
    _ensure_sheet(
        wb, config.comments_sheet,
        ["Prompt ID", "Comment", "Session ID", "Created At"],
    )
    _ensure_sheet(
        wb, config.history_sheet,
        ["Prompt ID", "Prompt Name", "Filled Values", "Session ID", "Created At"],
    )

    wb.save(wb_path)
    wb.close()


def _ensure_sheet(
    wb: openpyxl.Workbook, name: str, headers: List[str]
) -> None:
    if name not in wb.sheetnames:
        logger.info("Creating sheet '%s'", name)
        ws = wb.create_sheet(name)
        ws.append(headers)


# ── Load ──────────────────────────────────────────────────────────────────────

def load_feedback(
    path: str, config: AppConfig = APP_CONFIG
) -> Tuple[Dict[int, Dict], Dict[int, Dict]]:
    """
    Returns:
      ratings_map  {prompt_id: {"ratings": [int], "sessions": [str]}}
      comments_map {prompt_id: {"comments": [str], "timestamps": [str], "sessions": [str]}}
    """
    ratings_map: Dict[int, Dict]  = {}
    comments_map: Dict[int, Dict] = {}

    if not Path(path).exists():
        return ratings_map, comments_map

    ensure_feedback_file(path, config)

    # Ratings
    try:
        rdf = pd.read_excel(path, sheet_name=config.ratings_sheet)
        for _, row in rdf.iterrows():
            pid    = _safe_int(row.get("Prompt ID"))
            rating = _safe_int(row.get("Rating"))
            sess   = str(row.get("Session ID", "") or "")
            if pid and rating and 1 <= rating <= 5:
                b = ratings_map.setdefault(pid, {"ratings": [], "sessions": []})
                b["ratings"].append(rating)
                b["sessions"].append(sess)
        logger.info(
            "Ratings loaded: %d entries across %d prompts",
            sum(len(v["ratings"]) for v in ratings_map.values()),
            len(ratings_map),
        )
    except Exception as exc:
        logger.error("Failed to load ratings: %s", exc)

    # Comments
    try:
        cdf = pd.read_excel(path, sheet_name=config.comments_sheet)
        for _, row in cdf.iterrows():
            pid     = _safe_int(row.get("Prompt ID"))
            comment = str(row.get("Comment", "") or "").strip()
            ts      = str(row.get("Created At", "") or "")
            sess    = str(row.get("Session ID", "") or "")
            if pid and comment:
                b = comments_map.setdefault(
                    pid, {"comments": [], "timestamps": [], "sessions": []}
                )
                b["comments"].append(comment)
                b["timestamps"].append(ts)
                b["sessions"].append(sess)
        logger.info(
            "Comments loaded: %d entries across %d prompts",
            sum(len(v["comments"]) for v in comments_map.values()),
            len(comments_map),
        )
    except Exception as exc:
        logger.error("Failed to load comments: %s", exc)

    return ratings_map, comments_map


def load_history(
    path: str, config: AppConfig = APP_CONFIG
) -> List[Dict[str, Any]]:
    """Returns all saved filled-prompt history entries as a list of dicts."""
    if not Path(path).exists():
        return []
    try:
        df = pd.read_excel(path, sheet_name=config.history_sheet)
        rows = []
        for _, row in df.iterrows():
            try:
                filled = json.loads(str(row.get("Filled Values", "{}") or "{}"))
            except (json.JSONDecodeError, TypeError):
                filled = {}
            rows.append({
                "Prompt ID":   row.get("Prompt ID"),
                "Prompt Name": str(row.get("Prompt Name", "") or ""),
                "Filled Values": json.dumps(filled, ensure_ascii=False),
                "Session ID":  str(row.get("Session ID",  "") or ""),
                "Created At":  str(row.get("Created At",  "") or ""),
            })
        return rows
    except Exception as exc:
        logger.error("Failed to load history: %s", exc)
        return []


# ── Write ─────────────────────────────────────────────────────────────────────

def append_rating(
    path: str,
    prompt_id: int,
    rating: int,
    session_id: str = "",
    config: AppConfig = APP_CONFIG,
) -> None:
    _validate_id(prompt_id)
    if not (1 <= rating <= 5):
        raise ValueError(f"Rating must be 1–5, got {rating!r}")
    ensure_feedback_file(path, config)
    _append_row(path, config.ratings_sheet, [prompt_id, rating, session_id, _now()])
    logger.info("Rating saved: prompt_id=%d rating=%d", prompt_id, rating)


def append_comment(
    path: str,
    prompt_id: int,
    comment: str,
    session_id: str = "",
    config: AppConfig = APP_CONFIG,
) -> None:
    _validate_id(prompt_id)
    comment = comment.strip()
    if not comment:
        raise ValueError("Comment cannot be empty")
    if len(comment) > config.max_comment_length:
        raise ValueError(
            f"Comment exceeds {config.max_comment_length} characters (got {len(comment)})"
        )
    ensure_feedback_file(path, config)
    _append_row(path, config.comments_sheet, [prompt_id, comment, session_id, _now()])
    logger.info("Comment saved: prompt_id=%d", prompt_id)


def append_history(
    path: str,
    prompt_id: int,
    prompt_name: str,
    filled_values: Dict[str, str],
    session_id: str = "",
    config: AppConfig = APP_CONFIG,
) -> None:
    """Records which values a user filled into a prompt template."""
    _validate_id(prompt_id)
    ensure_feedback_file(path, config)
    _append_row(
        path,
        config.history_sheet,
        [prompt_id, prompt_name, json.dumps(filled_values, ensure_ascii=False), session_id, _now()],
    )
    logger.info("History saved: prompt_id=%d", prompt_id)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _append_row(path: str, sheet_name: str, values: List[Any]) -> None:
    wb = openpyxl.load_workbook(path)
    wb[sheet_name].append(values)
    wb.save(path)
    wb.close()


def _safe_int(value: Any) -> Optional[int]:
    try:
        return int(value) if pd.notna(value) else None
    except (TypeError, ValueError):
        return None


def _validate_id(prompt_id: int) -> None:
    if not isinstance(prompt_id, int) or prompt_id <= 0:
        raise ValueError(f"Invalid prompt_id: {prompt_id!r}")


def _now() -> str:
    return datetime.now().isoformat(timespec="seconds")
